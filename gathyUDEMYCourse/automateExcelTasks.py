
# WORKING WITH EXCEL THRU PYTHON
import openpyxl
workbook = openpyxl.load_workbook("/Users/jordanfelan/Downloads/Employees.xlsx")

# print(workbook.properties)
#
# print(workbook.sheetnames)
# print(workbook.active)

# sheet = workbook("EmployeeData")
# workbook.create_sheet("TestSheet")
# workbook.save("/Users/jordanfelan/Downloads/Employees.xlsx")

# sheet1 = workbook['TestSheet']
# sheet2 = workbook['TestSheet1']
# sheet3 = workbook['TestSheet2']
# workbook.remove(sheet1)
# workbook.remove(sheet2)
# workbook.remove(sheet3)
# workbook.save("/Users/jordanfelan/Downloads/Employees.xlsx")

sheet = workbook['EmployeeData']

print(sheet.title)
# print(dir(sheet))
#
# print(sheet.active_cell)
#
# print(sheet.dimensions)
#
# print(sheet.sheet_format)
# print(sheet.sheet_properties)
# print(sheet.sheet_state)
# print(sheet.sheet_view)
print(sheet.max_row)
print(sheet.max_column)

for i in sheet.values:
    print(i)

# WORKING WITH SHEET CELLS USING PYTHON

# HOW TO REFERENCE CERTAIN CELL
# print(sheet['B7'].value)
# print(sheet.cell(row = 6, column = 2).value)
#
# cell = sheet['B9']
#
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)
# print(cell.data_type)
# print(cell.encoding)
#
# cell = sheet['B2']
# cell.value = 'David'
# workbook.save("/Users/jordanfelan/Downloads/Employees.xlsx")
# print(cell.parent)

# Working with cell styles
from openpyxl.styles import *

cell = sheet['B8']

font = Font(color = 'FF0000', bold = True, italic = True)

cell.font = font

fill = PatternFill(fill_type = 'solid', bgColor = '000000')

cell.fill = fill

border = Border(left = Side(border_style = 'double', color = '322FEC'), right = Side(border_style = 'double', color = '322FEC'), top = Side(border_style = 'double', color = '322FEC'), bottom = Side(border_style = 'double', color = '322FEC'))

cell.border = border

align = Alignment(horizontal= 'left')

cell.alignment = align

workbook.save("/Users/jordanfelan/Downloads/Employees.xlsx")